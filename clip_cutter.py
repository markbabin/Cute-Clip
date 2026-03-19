import os
import sys
import re
import subprocess
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent.resolve()
INPUT_DIR = SCRIPT_DIR / "input"
OUTPUT_DIR = SCRIPT_DIR / "output"

# Channel name mapping: Excel name -> what appears in filenames
CHANNEL_ALIASES = {
    "planet tv":    ["planet tv hd", "planet tv"],
    "pop tv":       ["pop tv hd", "pop tv"],
    "rtv slo":      ["slo 1 hd", "slo1 hd", "rtv slo"],
    "rtv slo 1":    ["slo 1 hd", "slo1 hd"],
    "slo 1":        ["slo 1 hd", "slo1 hd"],
    "kanal a":      ["kanal a hd", "kanal a"],
}

# Shows on SLO 1 HD that are distinguished by broadcast time in the filename
# Dnevnik airs around 19:xx, Odmevi airs around 21:xx-22:xx
SLO1_SHOW_TIME_RANGES = {
    "dnevnik": (18, 20),   # filename time hour between 18-20
    "odmevi":  (20, 24),   # filename time hour between 20-24
}


def normalize(s):
    """Normalize a string for fuzzy comparison."""
    return re.sub(r'[\s_]+', ' ', s).strip().lower()


def parse_excel_date(val):
    """Parse Excel date value (could be datetime object or string like '1.1.26')."""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    # Handle d.m.yy or d.m.yyyy
    parts = s.split(".")
    if len(parts) == 3:
        day, month, year = parts
        day = int(day)
        month = int(month)
        year = int(year)
        if year < 100:
            year += 2000
        return f"{year:04d}-{month:02d}-{day:02d}"
    raise ValueError(f"Cannot parse date: {val}")


def parse_filename(fname):
    """Extract date, channel+show+time from a filename."""
    stem = Path(fname).stem
    # Date is always the first 10 characters: YYYY-MM-DD
    date_match = re.match(r'(\d{4}-\d{2}-\d{2})', stem)
    if not date_match:
        return None
    date_str = date_match.group(1)
    rest = stem[len(date_str):]
    # Strip leading separators
    rest = re.sub(r'^[\s_]+', '', rest)
    rest_normalized = normalize(rest)

    # Extract time from the end (HH-MM-SS pattern)
    time_match = re.search(r'(\d{1,2})-(\d{2})-(\d{2})$', rest)
    file_hour = int(time_match.group(1)) if time_match else None

    return {
        "stem": stem,
        "date": date_str,
        "rest": rest_normalized,
        "hour": file_hour,
    }


def find_matching_file(date_str, channel, show, mp4_files, parsed_cache):
    """Find the MP4 file matching a given date, channel, and show."""
    channel_norm = normalize(channel)
    show_norm = normalize(show)

    # Get possible channel name variants
    aliases = CHANNEL_ALIASES.get(channel_norm, [channel_norm])

    candidates = []
    for f in mp4_files:
        info = parsed_cache.get(f)
        if not info:
            continue
        if info["date"] != date_str:
            continue

        # Check if any channel alias appears in the filename
        channel_match = any(alias in info["rest"] for alias in aliases)
        if not channel_match:
            continue

        candidates.append((f, info))

    if not candidates:
        return None

    # If only one candidate, return it
    if len(candidates) == 1:
        return candidates[0][0]

    # Multiple candidates (e.g. Dnevnik vs Odmevi on SLO 1 HD)
    # Use time-based disambiguation
    time_range = SLO1_SHOW_TIME_RANGES.get(show_norm)
    if time_range:
        lo, hi = time_range
        for f, info in candidates:
            if info["hour"] is not None and lo <= info["hour"] < hi:
                return f

    # Fallback: try matching show name in filename
    for f, info in candidates:
        if show_norm in info["rest"]:
            return f

    # Last resort: return first match
    return candidates[0][0]


def sanitize_filename(s):
    """Make a string safe for use as a filename."""
    s = re.sub(r'[<>:"/\\|?*]', '', s)
    s = re.sub(r'\s+', '_', s.strip())
    return s


def format_timecode(val):
    """Ensure timecode is in HH:MM:SS format."""
    if isinstance(val, datetime):
        return val.strftime("%H:%M:%S")
    s = str(val).strip()
    # Already HH:MM:SS
    if re.match(r'^\d{1,2}:\d{2}:\d{2}$', s):
        return s
    # Handle H:MM:SS
    parts = s.split(":")
    if len(parts) == 3:
        return f"{int(parts[0]):02d}:{parts[1]}:{parts[2]}"
    return s


def cut_clip(input_path, output_path, start, end):
    """Use ffmpeg to cut a clip with fast copy."""
    cmd = [
        "ffmpeg", "-y",
        "-ss", start,
        "-to", end,
        "-i", str(input_path),
        "-c", "copy",
        "-avoid_negative_ts", "make_zero",
        str(output_path),
    ]
    print(f"  Running: ffmpeg -ss {start} -to {end} -i \"{input_path.name}\" -c copy \"{output_path.name}\"")
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"  ERROR: ffmpeg failed:\n{result.stderr[-500:]}")
        return False
    return True


def main():
    # Find Excel file
    if len(sys.argv) > 1:
        excel_path = Path(sys.argv[1])
    else:
        # Look for .xlsx in script directory
        xlsx_files = list(SCRIPT_DIR.glob("*.xlsx"))
        if not xlsx_files:
            print(f"No .xlsx file found in {SCRIPT_DIR}")
            print("Usage: python clip_cutter.py [spreadsheet.xlsx]")
            sys.exit(1)
        if len(xlsx_files) > 1:
            print("Multiple .xlsx files found, please specify one:")
            for f in xlsx_files:
                print(f"  {f.name}")
            sys.exit(1)
        excel_path = xlsx_files[0]

    print(f"Excel: {excel_path.name}")
    print(f"Input: {INPUT_DIR}")
    print(f"Output: {OUTPUT_DIR}")

    INPUT_DIR.mkdir(exist_ok=True)
    OUTPUT_DIR.mkdir(exist_ok=True)

    # Scan input folder for MP4 files
    mp4_files = list(INPUT_DIR.glob("*.mp4")) + list(INPUT_DIR.glob("*.MP4"))
    if not mp4_files:
        print(f"\nNo MP4 files found in {INPUT_DIR}")
        print("Place your MP4 files in the 'input' folder and run again.")
        sys.exit(1)

    print(f"Found {len(mp4_files)} MP4 file(s)\n")

    # Pre-parse all filenames
    parsed_cache = {}
    for f in mp4_files:
        parsed = parse_filename(f.name)
        if parsed:
            parsed_cache[f] = parsed

    # Read Excel
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    total = 0
    success = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not row or not row[0]:
            continue

        date_raw, channel, show, start_tc, end_tc, discipline = row[:6]

        # Skip rows with missing data
        if not all([date_raw, channel, show, start_tc, end_tc, discipline]):
            print(f"SKIP: Incomplete row: {row[:6]}")
            skipped += 1
            continue

        total += 1

        try:
            date_str = parse_excel_date(date_raw)
        except ValueError as e:
            print(f"SKIP: {e}")
            skipped += 1
            continue

        start = format_timecode(start_tc)
        end_ = format_timecode(end_tc)
        channel = str(channel).strip()
        show = str(show).strip()
        discipline = str(discipline).strip()

        # Find matching MP4
        match = find_matching_file(date_str, channel, show, mp4_files, parsed_cache)

        if not match:
            print(f"NO MATCH: {date_str} | {channel} | {show}")
            skipped += 1
            continue

        # Output filename: date-show-discipline.mp4
        out_name = f"{date_str}-{sanitize_filename(show)}-{sanitize_filename(discipline)}.mp4"
        out_path = OUTPUT_DIR / out_name

        # Handle duplicate names
        if out_path.exists():
            i = 2
            while out_path.exists():
                out_name = f"{date_str}-{sanitize_filename(show)}-{sanitize_filename(discipline)}_{i}.mp4"
                out_path = OUTPUT_DIR / out_name
                i += 1

        print(f"[{total}] {date_str} | {show} | {discipline}")
        print(f"  Source: {match.name}")
        print(f"  Cut: {start} -> {end_}")

        if cut_clip(match, out_path, start, end_):
            print(f"  OK: {out_name}")
            success += 1
        else:
            skipped += 1

        print()

    print(f"Done! {success}/{total} clips cut successfully. {skipped} skipped.")


if __name__ == "__main__":
    main()
